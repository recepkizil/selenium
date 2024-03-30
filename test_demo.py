from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains 
import pytest
import openpyxl
from constants.globalConstants import *

class Test_Demo:
    def deneme():
        print("deneme")

    #pytest tarafından tanımlanan bir method 
    #her test öncesi otomatik olarak çalıştırılır
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(BASE_URL)

    #her test bitiminde çalışacak fonk
    def teardown_method(self):
        self.driver.quit()

    @pytest.mark.skip #tüm testler koşulurken "skip" şeklinde işaretlenen testlerimi atla
    def test_demo(self):
        print("x")
        text = "Hello"
        assert text == "Hello"

    def getData():
        return [("1","1"),("abc","123"),("deneme","secret_sauce")]
    
    def readInvalidDataFromExcel():
        excelFile = openpyxl.load_workbook("data/invalidLogin.xlsx")
        sheet = excelFile["Sheet1"]
        rows = sheet.max_row #kaçıncı satıra kadar benim verim var
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data

    
    @pytest.mark.parametrize("username,password",readInvalidDataFromExcel())
    def test_invalid_login(self,username,password):
        #userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        userNameInput = self.waitForElementVisible((By.ID,username_id))
        #passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        passwordInput = self.waitForElementVisible((By.ID,password_id))
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        #loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        loginButton = self.waitForElementVisible((By.ID,login_button_id))
        loginButton.click()
        #errorMessage =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,errorMessage_xpath)))
        errorMessage = self.waitForElementVisible((By.ID,errorMessage_xpath))
        assert errorMessage.text == errorMessage_text

    def test_valid_login(self):
        self.driver.get("https://www.saucedemo.com/")
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"standard_user")
        actions.send_keys_to_element(passwordInput,"secret_sauce")
        actions.perform() #depoladığım aksiyonları çalıştır
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"login-button")))
        loginButton.click()
        baslik =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='header_container']/div[1]/div[2]/div")))
        assert baslik.text == "Swag Labs"
        
    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))