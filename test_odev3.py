from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait  #ilgili driverı bekleyen yapı
from selenium.webdriver.support import expected_conditions as ec  #
from selenium.webdriver.common.action_chains import ActionChains
import pytest
import openpyxl
from constants.globalConstants import * #klasör açtık ordan çekicez
# üstteki importu "from constants import globalConstants as c" şeklinde yazsaydık ordakileri her çağırışımızda başına "c."" yazacaktık. c.BASE_URL gibi
import json

class Test_Odev:

    def setup_method(self):
        self.driver=webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(BASE_URL) #constants klasörü açarak içine değişkenler oluşturduk ordan çektik

    def teardown_method(self):
        self.driver.quit()
    
    
    def test_blank_login(self):
        userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"")
        actions.send_keys_to_element(passwordInput,"")
        actions.click(loginButton)
        actions.perform() 
        errorMessage=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,blank_error_text_xpath)))
        assert errorMessage.text == blank_error_text
    
    def test_blank_password_login(self):
        userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"Recep")
        actions.send_keys_to_element(passwordInput,"")
        actions.click(loginButton)
        actions.perform() 
        errorMessage=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,blank_password_text_xpath)))
        assert errorMessage.text == blank_password_text
        

    def test_lockedUser_login(self):
        userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id))) 
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"locked_out_user")
        actions.send_keys_to_element(passwordInput,"secret_sauce")
        actions.click(loginButton)
        actions.perform() 
        errorMessage=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, locked_user_text_xpath)))
        assert errorMessage.text == locked_user_text

    def test_valid_login(self):
        userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"standard_user")
        actions.send_keys_to_element(passwordInput,"secret_sauce")
        actions.click(loginButton)
        actions.perform()
        baslik=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//div[@class='app_logo']")))
        assert baslik.text=="Swag Labs"
        listOfProducts=self.driver.find_elements(By.CSS_SELECTOR,"div[class='inventory_item']")
        print(len(listOfProducts))
        assert len(listOfProducts)==6

    def waitForElementVisible(self,locator,timeout=5): #visibility kısımlarını kısaltmak için onu da fonk yaptık. örnek olsun diye 98.satırda yaptım.
        return WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator)) #BURADAKİ RETURN OLMAZSA AŞAĞIDAKİ KISALTMALAR ÇALIŞMIYOR
        #webdriverwait ve ec ları döndürmek gerekiyor yoksa çalışmıyor.

    #def getData(): #bunu iptal ettik excelden çektik artık
        #return [("ahmet","12345"),("naber","5341231"),("Ahmet Suat Tanis","secret_sauce")]

    """def readInvalidDataFromExcel():
        excelFile = openpyxl.load_workbook("data/invalidLogin.xlsx") #dosyanın nerde olduğunu gösterdik data klasöründe 
        sheet = excelFile["Sheet1"] #sayfa değişkeni oluşturduk ve sayfayı söyledik
        rows = sheet.max_row #kaçıncı satıra kadar verim var onu söyledik
        data = []
        for i in range(2,rows+1): #parametreler 2.satırda olddan 2den başlattık, veri 4te bitiyor ama rows +1 yazıyoruz sondaki de dahil olsun diye
            username = sheet.cell(i,1).value #satırın 1.hücresi username'e gitsin. hücrenin içindeki değere ulaşmak için .value yazdık
            password = sheet.cell(i,2).value #satırın 2.hücresi password'e gitsin
            data.append((username,password)) 
        return data #kullanılan noktaya bu datayı göndermek istediğimizi söylüyoruz"""
    
    #artık pytest parametrize ile excel verilerimizi çağırabiliriz. get data örneğini siliyorum, parantez içine yeni defi yazıyorum
    #excelden sonra json öğrendik onu deniyoruz. import json yaptık

    def readInvalidDataFromJSON(json_file_path):
     with open(json_file_path, 'r') as file:
        data = json.load(file)
        invalid_users = data.get('invalid_login_users', [])
        return [(user.get('username'), user.get('password')) for user in invalid_users]

    @pytest.mark.parametrize("username,password",readInvalidDataFromJSON("invalid/data.json"))  
    def test_invalid_login(self,username,password):
        #userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id))) #eskiden böyleydi
        userNameInput = self.waitForElementVisible((By.ID,username_id)) 
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.click(loginButton)
        actions.perform()
        errorMessage=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,errorMessage_xpath)))
        assert errorMessage.text == errorMessage_text

    
    def test_addToCartProduct(self):
        userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"standard_user")
        actions.send_keys_to_element(passwordInput,"secret_sauce")
        actions.click(loginButton)
        actions.perform()
        item=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,"add-to-cart-sauce-labs-bike-light")))
        item.click()
        cartButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,".shopping_cart_badge")))
        cart_button_text=cartButton.text
        print("Cart Button Text:", cart_button_text)
        assert cartButton.text=="1"

    def test_removeProductFromCart(self):
        userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"standard_user")
        actions.send_keys_to_element(passwordInput,"secret_sauce")
        actions.click(loginButton)
        actions.perform()
        item1=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,"add-to-cart-sauce-labs-bike-light")))
        item1.click()
        item2=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"#add-to-cart-sauce-labs-backpack")))
        item2.click()
        cartButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//span[@class='shopping_cart_badge']")))
        cartButton.click()
        removeButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"#remove-sauce-labs-bike-light")))
        removeButton.click()
        cartButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//span[@class='shopping_cart_badge']")))
        assert cartButton.text=="1"

    def test_checkoutProduct(self):
        userNameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        loginButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"standard_user")
        
        actions.send_keys_to_element(passwordInput,"secret_sauce")
        
        actions.click(loginButton)
        actions.perform()
        item1=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,"add-to-cart-sauce-labs-bike-light")))
        item1.click()
        
        item2=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"#add-to-cart-sauce-labs-backpack")))
        item2.click()
        cartButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//span[@class='shopping_cart_badge']")))
        cartButton.click()
       
        checkoutButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"checkout")))
        checkoutButton.click()
        firstName=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"first-name")))
        lastName=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"last-name")))
        zipCode=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"postal-code")))
        continueButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"continue")))
        actions=ActionChains(self.driver)
        actions.send_keys_to_element(firstName,"Recep")
      
        actions.send_keys_to_element(lastName,"Kizil")
       
        actions.send_keys_to_element(zipCode,"34100")
       
        actions.perform()
        continueButton.click()
        finishButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"finish")))
        
        actions.click(finishButton)
        actions.perform()
        orderConfirmMessage=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,order_confirm_xpath)))
        assert orderConfirmMessage.text== order_confirm_message_text
    


    #BİR YER SARI VEYA BEYAZSA RETURN KULLANMAK SORUNU ÇÖZEBİLİR